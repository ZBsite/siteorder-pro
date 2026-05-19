import os
import sqlite3
import csv
import io
import json
import shutil
import re
from datetime import datetime
from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from werkzeug.utils import secure_filename
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import openpyxl
import pdfplumber

app = Flask(__name__)
CORS(app)

DATA_DIR = os.environ.get('RAILWAY_VOLUME_MOUNT_PATH', os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(DATA_DIR, 'construction_orders.db')
UPLOAD_DIR = os.path.join(DATA_DIR, 'uploads')

app.config['UPLOAD_FOLDER'] = UPLOAD_DIR
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

def get_db():
    from flask import g
    if not hasattr(g, '_db'):
        g._db = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
        g._db.row_factory = sqlite3.Row
        g._db.execute('PRAGMA journal_mode=WAL')
        g._db.execute('PRAGMA synchronous=NORMAL')
    return g._db

@app.teardown_appcontext
def close_db(error):
    from flask import g
    db = g.pop('_db', None)
    if db is not None:
        db.close()

def init_db():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA journal_mode=WAL')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS job_folders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        job_number TEXT,
        address TEXT,
        city TEXT,
        contact TEXT,
        phone TEXT,
        color TEXT DEFAULT \'#1e3a5f\',
        notes TEXT,
        created_at TEXT DEFAULT (datetime(\'now\'))
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_number TEXT UNIQUE NOT NULL,
        supplier_name TEXT NOT NULL,
        site_location TEXT,
        ordered_by TEXT,
        order_date TEXT NOT NULL,
        expected_delivery TEXT,
        status TEXT DEFAULT \'Pending\',
        notes TEXT,
        folder_id INTEGER REFERENCES job_folders(id),
        created_at TEXT DEFAULT (datetime(\'now\'))
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS order_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER NOT NULL,
        item_name TEXT NOT NULL,
        description TEXT,
        quantity_ordered REAL NOT NULL,
        quantity_received REAL DEFAULT 0,
        unit TEXT DEFAULT \'units\',
        unit_price REAL DEFAULT 0,
        status TEXT DEFAULT \'Pending\',
        delivery_note TEXT DEFAULT \'\',
        FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS deliveries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER NOT NULL,
        delivery_date TEXT NOT NULL,
        received_by TEXT,
        delivery_notes TEXT,
        photo_path TEXT,
        created_at TEXT DEFAULT (datetime(\'now\')),
        FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS order_attachments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER NOT NULL,
        filename TEXT NOT NULL,
        original_name TEXT NOT NULL,
        file_path TEXT NOT NULL,
        file_type TEXT,
        uploaded_at TEXT DEFAULT (datetime(\'now\')),
        notes TEXT,
        FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
    )''')
    # Migrations for existing DBs
    for col, defval in [('folder_id','INTEGER'), ('delivery_note','TEXT DEFAULT \'\'')]:
        try:
            if col == 'folder_id':
                c.execute('ALTER TABLE orders ADD COLUMN folder_id INTEGER REFERENCES job_folders(id)')
            else:
                c.execute(f'ALTER TABLE order_items ADD COLUMN {col} {defval}')
        except: pass
    conn.commit()
    conn.close()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.',1)[1].lower() in ALLOWED_EXTENSIONS

def generate_order_number(conn=None):
    if conn is None:
        conn = get_db()
    c = conn.cursor()
    year = datetime.now().year
    c.execute("SELECT order_number FROM orders WHERE order_number LIKE ? ORDER BY order_number DESC LIMIT 1", (f'ORD-{year}-%',))
    row = c.fetchone()
    if row:
        try: last_num = int(row[0].split('-')[-1])
        except: last_num = 0
        count = last_num + 1
    else:
        count = 1
    while True:
        candidate = f'ORD-{year}-{count:03d}'
        c.execute("SELECT 1 FROM orders WHERE order_number=?", (candidate,))
        if not c.fetchone():
            return candidate
        count += 1

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/dashboard')
def dashboard():
    conn = get_db()
    c = conn.cursor()
    stats = {}
    for s in ['Pending','Delivered','Partial','Backordered']:
        c.execute("SELECT COUNT(*) FROM orders WHERE status=?", (s,))
        stats[s.lower()] = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM orders")
    stats['total'] = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM order_items WHERE status='Backordered'")
    stats['backorder_items'] = c.fetchone()[0]
    c.execute("""SELECT o.order_number,o.supplier_name,o.site_location,o.expected_delivery,o.status
        FROM orders o WHERE o.status IN ('Pending','Partial','Backordered')
        ORDER BY o.expected_delivery ASC LIMIT 6""")
    stats['upcoming'] = [dict(r) for r in c.fetchall()]
    c.execute("SELECT supplier_name,COUNT(*) as cnt FROM orders GROUP BY supplier_name ORDER BY cnt DESC LIMIT 5")
    stats['by_supplier'] = [dict(r) for r in c.fetchall()]
    c.execute("SELECT jf.id,jf.name,jf.color,COUNT(o.id) as order_count FROM job_folders jf LEFT JOIN orders o ON o.folder_id=jf.id GROUP BY jf.id ORDER BY jf.name")
    stats['folders'] = [dict(r) for r in c.fetchall()]
    return jsonify(stats)

@app.route('/api/orders', methods=['GET'])
def get_orders():
    conn = get_db()
    c = conn.cursor()
    status = request.args.get('status','')
    search = request.args.get('search','')
    site = request.args.get('site','')
    folder_id = request.args.get('folder_id','')
    q = "SELECT o.*, jf.name as folder_name, jf.color as folder_color FROM orders o LEFT JOIN job_folders jf ON o.folder_id=jf.id WHERE 1=1"
    p = []
    if status: q += " AND o.status=?"; p.append(status)
    if search:
        q += " AND (o.order_number LIKE ? OR o.supplier_name LIKE ? OR o.ordered_by LIKE ?)"
        p += [f'%{search}%',f'%{search}%',f'%{search}%']
    if site: q += " AND o.site_location LIKE ?"; p.append(f'%{site}%')
    if folder_id: q += " AND o.folder_id=?"; p.append(folder_id)
    q += " ORDER BY o.created_at DESC"
    c.execute(q, p)
    orders = [dict(r) for r in c.fetchall()]
    for order in orders:
        c.execute("SELECT * FROM order_items WHERE order_id=?", (order['id'],))
        order['items'] = [dict(i) for i in c.fetchall()]
        c.execute("SELECT * FROM deliveries WHERE order_id=? ORDER BY delivery_date DESC", (order['id'],))
        order['deliveries'] = [dict(d) for d in c.fetchall()]
        c.execute("SELECT * FROM order_attachments WHERE order_id=? ORDER BY uploaded_at DESC", (order['id'],))
        order['attachments'] = [dict(a) for a in c.fetchall()]
    return jsonify(orders)

@app.route('/api/orders', methods=['POST'])
def create_order():
    data = request.json
    conn = get_db()
    c = conn.cursor()
    order_number = generate_order_number(conn)
    folder_id = data.get('folder_id') or None
    c.execute('''INSERT INTO orders (order_number,supplier_name,site_location,ordered_by,order_date,expected_delivery,status,notes,folder_id)
                 VALUES (?,?,?,?,?,?,?,?,?)''',
              (order_number, data['supplier_name'], data.get('site_location',''),
               data.get('ordered_by',''), data['order_date'],
               data.get('expected_delivery',''), 'Pending', data.get('notes',''), folder_id))
    order_id = c.lastrowid
    for item in data.get('items',[]):
        c.execute('''INSERT INTO order_items (order_id,item_name,description,quantity_ordered,quantity_received,unit,unit_price,status)
                     VALUES (?,?,?,?,0,?,?,?)''',
                  (order_id, item['item_name'], item.get('description',''),
                   item['quantity_ordered'], item.get('unit','units'), item.get('unit_price',0), 'Pending'))
    conn.commit()
    return jsonify({'success':True,'order_number':order_number,'order_id':order_id})

@app.route('/api/orders/<int:order_id>', methods=['GET'])
def get_order(order_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM orders WHERE id=?", (order_id,))
    order = c.fetchone()
    if not order: return jsonify({'error':'Not found'}), 404
    order = dict(order)
    c.execute("SELECT * FROM order_items WHERE order_id=?", (order_id,))
    order['items'] = [dict(i) for i in c.fetchall()]
    c.execute("SELECT * FROM deliveries WHERE order_id=? ORDER BY delivery_date DESC", (order_id,))
    order['deliveries'] = [dict(d) for d in c.fetchall()]
    c.execute("SELECT * FROM order_attachments WHERE order_id=? ORDER BY uploaded_at DESC", (order_id,))
    order['attachments'] = [dict(a) for a in c.fetchall()]
    return jsonify(order)

@app.route('/api/orders/<int:order_id>', methods=['PUT'])
def update_order(order_id):
    data = request.json
    conn = get_db()
    c = conn.cursor()
    folder_id = data.get('folder_id') or None
    c.execute('''UPDATE orders SET supplier_name=?,site_location=?,ordered_by=?,
                 order_date=?,expected_delivery=?,status=?,notes=?,folder_id=? WHERE id=?''',
              (data['supplier_name'],data.get('site_location',''),data.get('ordered_by',''),
               data['order_date'],data.get('expected_delivery',''),data.get('status','Pending'),
               data.get('notes',''),folder_id,order_id))
    conn.commit()
    return jsonify({'success':True})

@app.route('/api/orders/<int:order_id>', methods=['DELETE'])
def delete_order(order_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM orders WHERE id=?", (order_id,))
    conn.commit()
    return jsonify({'success':True})

@app.route('/api/orders/<int:order_id>/deliver', methods=['POST'])
def record_delivery(order_id):
    conn = get_db()
    c = conn.cursor()
    photo_path = None
    if 'photo' in request.files:
        file = request.files['photo']
        if file and allowed_file(file.filename):
            filename = secure_filename(f"{order_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
            photo_dir = os.path.join(UPLOAD_DIR, 'photos')
            os.makedirs(photo_dir, exist_ok=True)
            file.save(os.path.join(photo_dir, filename))
            photo_path = f"uploads/photos/{filename}"
    c.execute('''INSERT INTO deliveries (order_id,delivery_date,received_by,delivery_notes,photo_path)
                 VALUES (?,?,?,?,?)''',
              (order_id, request.form.get('delivery_date', datetime.now().strftime('%Y-%m-%d')),
               request.form.get('received_by',''), request.form.get('delivery_notes',''), photo_path))
    items = json.loads(request.form.get('items_json','[]'))
    for item in items:
        c.execute("SELECT quantity_ordered,quantity_received FROM order_items WHERE id=?", (item['id'],))
        row = c.fetchone()
        if row:
            new_recv = min(float(row['quantity_received']) + float(item.get('qty_received',0)), float(row['quantity_ordered']))
            if new_recv >= float(row['quantity_ordered']): istatus = 'Delivered'
            elif new_recv > 0: istatus = 'Partial'
            else: istatus = 'Backordered'
            item_note = item.get('note', '')
            c.execute("UPDATE order_items SET quantity_received=?,status=?,delivery_note=? WHERE id=?", (new_recv, istatus, item_note, item['id']))
    c.execute("SELECT status FROM order_items WHERE order_id=?", (order_id,))
    statuses = [r[0] for r in c.fetchall()]
    if all(s=='Delivered' for s in statuses): order_status = 'Delivered'
    elif any(s in ('Delivered','Partial') for s in statuses): order_status = 'Partial'
    elif any(s=='Backordered' for s in statuses): order_status = 'Backordered'
    else: order_status = 'Pending'
    c.execute("UPDATE orders SET status=? WHERE id=?", (order_status, order_id))
    conn.commit()
    return jsonify({'success':True,'order_status':order_status})

@app.route('/api/backorders')
def get_backorders():
    conn = get_db()
    c = conn.cursor()
    c.execute('''SELECT oi.*,o.order_number,o.supplier_name,o.site_location,o.expected_delivery,o.ordered_by
        FROM order_items oi JOIN orders o ON oi.order_id=o.id
        WHERE oi.status IN (\'Backordered\',\'Partial\',\'Pending\')
        ORDER BY o.expected_delivery ASC''')
    items = [dict(r) for r in c.fetchall()]
    return jsonify(items)

# --- JOB FOLDERS ---
@app.route('/api/folders', methods=['GET'])
def get_folders():
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT jf.*, COUNT(o.id) as order_count FROM job_folders jf
                 LEFT JOIN orders o ON o.folder_id=jf.id GROUP BY jf.id ORDER BY jf.name""")
    return jsonify([dict(r) for r in c.fetchall()])

@app.route('/api/folders', methods=['POST'])
def create_folder():
    data = request.json
    if not data.get('name'): return jsonify({'error':'Folder name is required'}), 400
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute('INSERT INTO job_folders (name,job_number,address,city,contact,phone,color,notes) VALUES (?,?,?,?,?,?,?,?)',
                  (data['name'],data.get('job_number',''),data.get('address',''),data.get('city',''),
                   data.get('contact',''),data.get('phone',''),data.get('color','#1e3a5f'),data.get('notes','')))
        folder_id = c.lastrowid
        conn.commit()
        return jsonify({'success':True,'id':folder_id})
    except Exception as e:
        return jsonify({'error':str(e)}), 400

@app.route('/api/folders/<int:folder_id>', methods=['PUT'])
def update_folder(folder_id):
    data = request.json
    conn = get_db()
    c = conn.cursor()
    c.execute('UPDATE job_folders SET name=?,job_number=?,address=?,city=?,contact=?,phone=?,color=?,notes=? WHERE id=?',
              (data['name'],data.get('job_number',''),data.get('address',''),data.get('city',''),
               data.get('contact',''),data.get('phone',''),data.get('color','#1e3a5f'),data.get('notes',''),folder_id))
    conn.commit()
    return jsonify({'success':True})

@app.route('/api/folders/<int:folder_id>', methods=['DELETE'])
def delete_folder(folder_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('UPDATE orders SET folder_id=NULL WHERE folder_id=?', (folder_id,))
    c.execute('DELETE FROM job_folders WHERE id=?', (folder_id,))
    conn.commit()
    return jsonify({'success':True})

@app.route('/api/folders/<int:folder_id>/orders', methods=['GET'])
def get_folder_orders(folder_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT o.*,jf.name as folder_name,jf.color as folder_color FROM orders o LEFT JOIN job_folders jf ON o.folder_id=jf.id WHERE o.folder_id=? ORDER BY o.created_at DESC", (folder_id,))
    orders = [dict(r) for r in c.fetchall()]
    for order in orders:
        c.execute("SELECT * FROM order_items WHERE order_id=?", (order['id'],))
        order['items'] = [dict(i) for i in c.fetchall()]
    return jsonify(orders)

# --- ATTACHMENTS & MERGE ---
@app.route('/api/orders/<int:order_id>/attachments', methods=['POST'])
def upload_attachment(order_id):
    if 'file' not in request.files: return jsonify({'error':'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename: return jsonify({'error':'No file selected'}), 400
    ext = file.filename.rsplit('.',1)[-1].lower() if '.' in file.filename else ''
    allowed = {'xlsx','xls','pdf','png','jpg','jpeg','gif','webp','csv','doc','docx','txt'}
    if ext not in allowed: return jsonify({'error':'File type not supported'}), 400
    att_dir = os.path.join(UPLOAD_DIR, 'attachments')
    os.makedirs(att_dir, exist_ok=True)
    stored_name = str(order_id)+'_'+datetime.now().strftime('%Y%m%d%H%M%S')+'_'+secure_filename(file.filename)
    file_path = os.path.join(att_dir, stored_name)
    file.save(file_path)
    conn = get_db()
    c = conn.cursor()
    c.execute('INSERT INTO order_attachments (order_id,filename,original_name,file_path,file_type,notes) VALUES (?,?,?,?,?,?)',
              (order_id,stored_name,file.filename,file_path,ext,request.form.get('notes','')))
    att_id = c.lastrowid
    conn.commit()
    return jsonify({'success':True,'id':att_id,'original_name':file.filename,'file_path':file_path,'file_type':ext})

@app.route('/api/orders/<int:order_id>/attachments/<int:att_id>', methods=['DELETE'])
def delete_attachment(order_id, att_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT file_path FROM order_attachments WHERE id=? AND order_id=?', (att_id,order_id))
    row = c.fetchone()
    if row:
        try: os.remove(row['file_path'])
        except: pass
        c.execute('DELETE FROM order_attachments WHERE id=?', (att_id,))
        conn.commit()
    return jsonify({'success':True})

@app.route('/api/orders/<int:order_id>/merge-items/preview', methods=['POST'])
def merge_items_preview(order_id):
    if 'file' not in request.files: return jsonify({'error':'No file uploaded'}), 400
    file = request.files['file']
    ext = file.filename.rsplit('.',1)[-1].lower() if '.' in file.filename else ''
    if ext not in ('xlsx','xls','pdf'): return jsonify({'error':'Only .xlsx, .xls, and .pdf supported'}), 400
    imp_dir = os.path.join(UPLOAD_DIR, 'imports')
    os.makedirs(imp_dir, exist_ok=True)
    tmp_path = os.path.join(imp_dir, f'prev_{order_id}_{datetime.now().strftime("%Y%m%d%H%M%S")}.{ext}')
    file.save(tmp_path)
    try:
        parsed, err = parse_excel_file(tmp_path) if ext in ('xlsx','xls') else parse_pdf_file(tmp_path)
        if err: return jsonify({'error':err}), 400
        all_items = [item for o in (parsed or []) for item in o.get('items',[])]
        return jsonify({'items':all_items,'count':len(all_items),'tmp_file':tmp_path})
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/api/orders/<int:order_id>/merge-items', methods=['POST'])
def merge_items(order_id):
    if 'file' not in request.files: return jsonify({'error':'No file uploaded'}), 400
    file = request.files['file']
    ext = file.filename.rsplit('.',1)[-1].lower() if '.' in file.filename else ''
    if ext not in ('xlsx','xls','pdf'): return jsonify({'error':'Only .xlsx, .xls, and .pdf supported'}), 400
    imp_dir = os.path.join(UPLOAD_DIR, 'imports')
    att_dir = os.path.join(UPLOAD_DIR, 'attachments')
    os.makedirs(imp_dir, exist_ok=True)
    os.makedirs(att_dir, exist_ok=True)
    tmp_path = os.path.join(imp_dir, f'merge_{order_id}_{datetime.now().strftime("%Y%m%d%H%M%S")}.{ext}')
    file.save(tmp_path)
    stored_name = str(order_id)+'_'+datetime.now().strftime('%Y%m%d%H%M%S')+'_'+secure_filename(file.filename)
    att_path = os.path.join(att_dir, stored_name)
    shutil.copy(tmp_path, att_path)
    try:
        parsed, err = parse_excel_file(tmp_path) if ext in ('xlsx','xls') else parse_pdf_file(tmp_path)
        if err: return jsonify({'error':err}), 400
        all_items = [item for o in (parsed or []) for item in o.get('items',[])]
        if not all_items: return jsonify({'error':'No line items found'}), 400
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT id FROM orders WHERE id=?', (order_id,))
        if not c.fetchone(): return jsonify({'error':'Order not found'}), 404
        for item in all_items:
            c.execute('INSERT INTO order_items (order_id,item_name,description,quantity_ordered,quantity_received,unit,unit_price,status) VALUES (?,?,?,?,0,?,?,?)',
                      (order_id,item['item_name'],item.get('description',''),item['quantity_ordered'],item.get('unit','units'),item.get('unit_price',0),'Pending'))
        c.execute('INSERT INTO order_attachments (order_id,filename,original_name,file_path,file_type,notes) VALUES (?,?,?,?,?,?)',
                  (order_id,stored_name,file.filename,att_path,ext,f'Imported from {file.filename} - {len(all_items)} items merged'))
        conn.commit()
        return jsonify({'success':True,'items_added':len(all_items)})
    except Exception as e:
        return jsonify({'error':str(e)}), 500

# --- EXPORT ---
@app.route('/api/export/csv')
def export_csv():
    conn = get_db()
    c = conn.cursor()
    status = request.args.get('status','')
    q = "SELECT * FROM orders"
    p = []
    if status: q += " WHERE status=?"; p.append(status)
    q += " ORDER BY created_at DESC"
    c.execute(q, p)
    orders = [dict(r) for r in c.fetchall()]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Order Number','Supplier','Site Location','Ordered By','Order Date','Expected Delivery','Status','Notes'])
    for o in orders:
        writer.writerow([o['order_number'],o['supplier_name'],o['site_location'],o['ordered_by'],o['order_date'],o['expected_delivery'],o['status'],o['notes']])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv',
                     as_attachment=True, download_name=f'orders_{datetime.now().strftime("%Y%m%d")}.csv')

@app.route('/api/export/pdf')
def export_pdf():
    conn = get_db()
    c = conn.cursor()
    status = request.args.get('status','')
    q = "SELECT * FROM orders"
    p = []
    if status: q += " WHERE status=?"; p.append(status)
    q += " ORDER BY created_at DESC"
    c.execute(q, p)
    orders = [dict(r) for r in c.fetchall()]
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph("<b>Construction Site Orders Report</b>", styles['Title']),
                Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Total: {len(orders)}", styles['Normal']),
                Spacer(1, 0.4*cm)]
    data = [['Order #','Supplier','Site','Ordered By','Order Date','Exp. Delivery','Status']]
    status_colors = {'Delivered':colors.HexColor('#d1fae5'),'Pending':colors.HexColor('#fef3c7'),
                     'Partial':colors.HexColor('#dbeafe'),'Backordered':colors.HexColor('#fee2e2')}
    for o in orders:
        data.append([o['order_number'],o['supplier_name'],o['site_location'] or '',o['ordered_by'] or '',o['order_date'],o['expected_delivery'] or '',o['status']])
    col_widths = [3.2*cm,4.5*cm,4.5*cm,3.5*cm,2.8*cm,3.2*cm,2.8*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,0),9),
        ('FONTSIZE',(0,1),(-1,-1),8),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f8fafc')]),
        ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#e2e8f0')),
        ('ALIGN',(0,0),(-1,-1),'LEFT'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('PADDING',(0,0),(-1,-1),5),
    ])
    for i,o in enumerate(orders,1):
        bg = status_colors.get(o['status'],colors.white)
        style.add('BACKGROUND',(6,i),(6,i),bg)
    table.setStyle(style)
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True,
                     download_name=f'orders_{datetime.now().strftime("%Y%m%d")}.pdf')

# --- IMPORT ---
def is_mtech_form(wb):
    for shname in wb.sheetnames:
        ws = wb[shname]
        cell = ws['A1'].value
        if cell and 'MTECH' in str(cell).upper() and 'MATERIAL REQUISITION' in str(cell).upper():
            return True
    return False

def parse_mtech_form(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    orders = []
    def clean(val, default=''):
        if val is None: return default
        s = str(val).strip()
        if s in ('0','1899-12-30','00:00:00','1899-12-30 00:00:00'): return default
        return s
    def parse_date(val):
        if val is None: return ''
        if hasattr(val, 'strftime'): return val.strftime('%Y-%m-%d')
        s = str(val).strip()
        if s in ('0','1899-12-30','00:00:00',''): return ''
        for fmt in ['%Y-%m-%d','%d/%m/%Y','%m/%d/%Y','%d-%m-%Y']:
            try: return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
            except: pass
        return s
    for shname in wb.sheetnames:
        if not shname.strip().upper().startswith('PG'): continue
        ws = wb[shname]
        order_date = parse_date(ws['C2'].value)
        job_name = clean(ws['C3'].value)
        job_number = clean(ws['H3'].value)
        job_address = clean(ws['C4'].value)
        city = clean(ws['H4'].value)
        deliver_to = clean(ws['C5'].value)
        delivery_date = parse_date(ws['I5'].value)
        contact = clean(ws['C6'].value)
        phone = clean(ws['I6'].value)
        tag = clean(ws['B7'].value)
        level = clean(ws['H7'].value)
        supplier_name = deliver_to if deliver_to else (job_name if job_name else 'MTECH Requisition')
        site_parts = [p for p in [job_name, job_address, city] if p]
        site_location = ' | '.join(site_parts)
        notes_parts = []
        if job_number: notes_parts.append(f'Job#: {job_number}')
        if tag: notes_parts.append(f'Tag: {tag}')
        if level: notes_parts.append(f'Level: {level}')
        if phone: notes_parts.append(f'Contact Phone: {phone}')
        notes = ' | '.join(notes_parts)
        items = []
        for row in ws.iter_rows(min_row=9, max_row=200, values_only=True):
            phase = clean(row[1] if len(row)>1 else None)
            qty_raw = row[2] if len(row)>2 else None
            description = clean(row[4] if len(row)>4 else None)
            brand = clean(row[9] if len(row)>9 else None)
            model = clean(row[10] if len(row)>10 else None)
            if not description and not qty_raw: continue
            if not description: continue
            try: qty = float(qty_raw) if qty_raw not in (None,'',0) else 1.0
            except: qty = 1.0
            desc_parts = []
            if phase: desc_parts.append(f'Phase: {phase}')
            if brand: desc_parts.append(f'Brand: {brand}')
            if model: desc_parts.append(f'Model: {model}')
            items.append({'item_name':description,'description':' | '.join(desc_parts),'quantity_ordered':qty,'unit':'units','unit_price':0.0})
        if items or (job_name or deliver_to or order_date):
            if not order_date: order_date = datetime.now().strftime('%Y-%m-%d')
            orders.append({'supplier_name':supplier_name or 'MTECH Requisition','site_location':site_location,
                           'ordered_by':contact,'order_date':order_date,'expected_delivery':delivery_date,
                           'notes':notes,'items':items,'_sheet':shname})
    orders = [o for o in orders if o['items'] or o['site_location'] or o['ordered_by']]
    if not orders:
        return [], 'MTECH form detected but no data found. Please fill in the form and re-upload.'
    return orders, None

def parse_excel_file(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if is_mtech_form(wb):
        return parse_mtech_form(filepath)
    rows = list(wb.active.iter_rows(values_only=True))
    if not rows: return [], "Empty spreadsheet"
    header_row = None
    header_idx = 0
    for i, row in enumerate(rows):
        non_empty = [str(c).strip().lower() for c in row if c is not None]
        if len(non_empty) >= 2:
            header_row = non_empty; header_idx = i; break
    if header_row is None: return [], "Could not find header row"
    col_map = {}
    for j, h in enumerate(header_row):
        if any(x in h for x in ['supplier','vendor','company']): col_map['supplier_name'] = j
        elif any(x in h for x in ['site','location','project']): col_map['site_location'] = j
        elif any(x in h for x in ['ordered by','ordered_by','requester']): col_map['ordered_by'] = j
        elif any(x in h for x in ['order date','order_date','date ordered','date']): col_map['order_date'] = j
        elif any(x in h for x in ['delivery','expected','due date','due']): col_map['expected_delivery'] = j
        elif any(x in h for x in ['item','material','product','description','name']): col_map['item_name'] = j
        elif any(x in h for x in ['qty','quantity','amount','ordered']): col_map['quantity_ordered'] = j
        elif any(x in h for x in ['unit price','price','cost','rate']): col_map['unit_price'] = j
        elif any(x in h for x in ['unit','uom','measure']): col_map['unit'] = j
        elif any(x in h for x in ['note','comment','remark']): col_map['notes'] = j
    orders_dict = {}
    for i, row in enumerate(rows[header_idx+1:], start=header_idx+2):
        if all(c is None or str(c).strip()=='' for c in row): continue
        def get(key, default=''):
            idx = col_map.get(key)
            if idx is None: return default
            val = row[idx] if idx < len(row) else None
            return str(val).strip() if val is not None else default
        supplier = get('supplier_name')
        item_name = get('item_name')
        if not supplier and not item_name: continue
        order_date = get('order_date', '')
        if order_date:
            for fmt in ['%Y-%m-%d','%d/%m/%Y','%m/%d/%Y','%d-%m-%Y']:
                try: order_date = datetime.strptime(order_date, fmt).strftime('%Y-%m-%d'); break
                except: pass
        if not order_date: order_date = datetime.now().strftime('%Y-%m-%d')
        exp_delivery = get('expected_delivery', '')
        if exp_delivery:
            for fmt in ['%Y-%m-%d','%d/%m/%Y','%m/%d/%Y','%d-%m-%Y']:
                try: exp_delivery = datetime.strptime(exp_delivery, fmt).strftime('%Y-%m-%d'); break
                except: pass
        key = f"{supplier}|{order_date}"
        if key not in orders_dict:
            orders_dict[key] = {'supplier_name':supplier or 'Unknown Supplier','site_location':get('site_location'),
                                'ordered_by':get('ordered_by'),'order_date':order_date,
                                'expected_delivery':exp_delivery,'notes':get('notes'),'items':[]}
        try: qty = float(get('quantity_ordered',1) or 1)
        except: qty = 1
        try: price = float(get('unit_price',0) or 0)
        except: price = 0
        if item_name:
            orders_dict[key]['items'].append({'item_name':item_name,'description':'','quantity_ordered':qty,
                                               'unit':get('unit','units') or 'units','unit_price':price})
    return list(orders_dict.values()), None

def parse_pdf_file(filepath):
    orders = []
    with pdfplumber.open(filepath) as pdf:
        all_text = ''
        all_tables = []
        for page in pdf.pages:
            all_text += page.extract_text() or ''
            tables = page.extract_tables()
            if tables: all_tables.extend(tables)
    if all_tables:
        for table in all_tables:
            if not table or len(table) < 2: continue
            header = [str(c).strip().lower() if c else '' for c in table[0]]
            col_map = {}
            for j, h in enumerate(header):
                if any(x in h for x in ['supplier','vendor']): col_map['supplier_name'] = j
                elif any(x in h for x in ['item','material','product','description']): col_map['item_name'] = j
                elif any(x in h for x in ['qty','quantity','amount']): col_map['quantity_ordered'] = j
                elif any(x in h for x in ['unit price','price','cost']): col_map['unit_price'] = j
                elif any(x in h for x in ['unit','uom']): col_map['unit'] = j
                elif any(x in h for x in ['date']): col_map['order_date'] = j
                elif any(x in h for x in ['delivery','due']): col_map['expected_delivery'] = j
            if 'item_name' not in col_map and 'supplier_name' not in col_map: continue
            orders_dict = {}
            for row in table[1:]:
                if not row or all(c is None or str(c).strip()=='' for c in row): continue
                def gcol(key, default=''):
                    idx = col_map.get(key)
                    if idx is None: return default
                    val = row[idx] if idx < len(row) else None
                    return str(val).strip() if val is not None else default
                supplier = gcol('supplier_name', 'Imported Supplier')
                item_name = gcol('item_name')
                order_date = gcol('order_date', datetime.now().strftime('%Y-%m-%d'))
                key = f"{supplier}|{order_date}"
                if key not in orders_dict:
                    orders_dict[key] = {'supplier_name':supplier,'site_location':gcol('site_location'),
                                        'ordered_by':'','order_date':order_date,'expected_delivery':gcol('expected_delivery'),
                                        'notes':'Imported from PDF','items':[]}
                try: qty = float(gcol('quantity_ordered',1) or 1)
                except: qty = 1
                try: price = float(re.sub(r'[^\d.]','',gcol('unit_price','0')) or 0)
                except: price = 0
                if item_name:
                    orders_dict[key]['items'].append({'item_name':item_name,'description':'','quantity_ordered':qty,'unit':gcol('unit','units') or 'units','unit_price':price})
            orders.extend(list(orders_dict.values()))
    if not orders and all_text:
        lines = [l.strip() for l in all_text.split('\n') if l.strip()]
        supplier = 'Imported Supplier'
        order_date = datetime.now().strftime('%Y-%m-%d')
        items = []
        for line in lines:
            sup_match = re.search(r'(?:supplier|vendor|from)[:\s]+(.+)', line, re.I)
            if sup_match: supplier = sup_match.group(1).strip()
            qty_match = re.search(r'(\d+(?:\.\d+)?)\s*(?:x|units?|bags?|pcs?|pieces?|tonnes?|m3|kg|sheets?|lengths?)\s+(.+)', line, re.I)
            if qty_match:
                items.append({'item_name':qty_match.group(2).strip(),'description':'','quantity_ordered':float(qty_match.group(1)),'unit':'units','unit_price':0})
        if items:
            orders.append({'supplier_name':supplier,'site_location':'','ordered_by':'','order_date':order_date,'expected_delivery':'','notes':'Imported from PDF','items':items})
    return orders, None

def save_imported_orders(orders_data):
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA journal_mode=WAL')
    c = conn.cursor()
    created = 0; skipped = 0
    for order in orders_data:
        if not order.get('supplier_name') or not order.get('items'): skipped += 1; continue
        order_number = generate_order_number(conn)
        c.execute('INSERT INTO orders (order_number,supplier_name,site_location,ordered_by,order_date,expected_delivery,status,notes) VALUES (?,?,?,?,?,?,?,?)',
                  (order_number,order['supplier_name'],order.get('site_location',''),order.get('ordered_by',''),
                   order['order_date'],order.get('expected_delivery',''),'Pending',order.get('notes','')))
        order_id = c.lastrowid
        for item in order['items']:
            c.execute('INSERT INTO order_items (order_id,item_name,description,quantity_ordered,quantity_received,unit,unit_price,status) VALUES (?,?,?,?,0,?,?,?)',
                      (order_id,item['item_name'],item.get('description',''),item['quantity_ordered'],item.get('unit','units'),item.get('unit_price',0),'Pending'))
        created += 1
    conn.commit(); conn.close()
    return created, skipped

@app.route('/api/import/preview', methods=['POST'])
def import_preview():
    if 'file' not in request.files: return jsonify({'error':'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename: return jsonify({'error':'No file selected'}), 400
    ext = file.filename.rsplit('.',1)[-1].lower()
    if ext not in ('xlsx','xls','pdf'): return jsonify({'error':'Only .xlsx, .xls, and .pdf files are supported'}), 400
    imp_dir = os.path.join(UPLOAD_DIR, 'imports')
    os.makedirs(imp_dir, exist_ok=True)
    tmp_path = os.path.join(imp_dir, f'tmp_import_{datetime.now().strftime("%Y%m%d%H%M%S")}.{ext}')
    file.save(tmp_path)
    try:
        orders, err = parse_excel_file(tmp_path) if ext in ('xlsx','xls') else parse_pdf_file(tmp_path)
        if err: return jsonify({'error':err}), 400
        return jsonify({'orders':orders,'count':len(orders),'total_items':sum(len(o['items']) for o in orders),'tmp_file':tmp_path,'file_type':ext})
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/api/import/confirm', methods=['POST'])
def import_confirm():
    data = request.json
    orders = data.get('orders',[])
    if not orders: return jsonify({'error':'No orders to import'}), 400
    created, skipped = save_imported_orders(orders)
    return jsonify({'success':True,'created':created,'skipped':skipped})

@app.route('/api/import/template/excel')
def download_excel_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    headers = ['Supplier Name','Site Location','Ordered By','Order Date','Expected Delivery','Item Name','Quantity','Unit','Unit Price','Notes']
    from openpyxl.styles import Font, PatternFill, Alignment
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    samples = [
        ['BuildRight Supplies','Site A - Main St','Mike Johnson','2024-06-01','2024-06-08','Portland Cement',100,'bags',12.50,'Urgent'],
        ['BuildRight Supplies','Site A - Main St','Mike Johnson','2024-06-01','2024-06-08','Sand',20,'tonnes',45.00,''],
        ['SteelCo Materials','Site B - Oak Ave','Sarah Lee','2024-06-03','2024-06-10','Steel Rebar 12mm',200,'pieces',8.75,''],
    ]
    for row_data in samples:
        ws.append(row_data)
    col_widths = [22,20,16,14,18,22,10,10,12,20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='order_import_template.xlsx')

if __name__ == '__main__':
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_DIR, 'attachments'), exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_DIR, 'imports'), exist_ok=True)
    init_db()
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') == 'development'
    app.run(host='0.0.0.0', port=port, debug=debug)
